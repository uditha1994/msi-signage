from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import json, os, uuid, socket
from datetime import datetime
from werkzeug.utils import secure_filename

# ── Load .env (Firebase config + other secrets) ─────────────
def load_env_file(path='.env'):
    if not os.path.exists(path): return
    with open(path) as f:
        for line in f:
            line=line.strip()
            if not line or line.startswith('#') or '=' not in line: continue
            k,v=line.split('=',1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
load_env_file()

app = Flask(__name__, static_folder='public')
CORS(app)

DATA_FILE      = 'timetable_data.json'
PRESETS_FILE   = 'presets.json'
TEMPLATES_FILE = 'templates_data.json'
CATALOG_FILE   = 'catalog.json'        # programs -> modules -> lessons
LECTURERS_FILE = 'lecturers.json'      # lecturer profiles with images
MEETING_FILE   = 'meeting.json'        # special meeting mode
MEDIA_FOLDER   = 'public/media'
PROFILE_FOLDER = 'public/profiles'
MEETING_FOLDER = 'public/meeting'
ALLOWED_EXT    = {'png','jpg','jpeg','gif','mp4','webm','mov'}
IMG_EXT        = {'png','jpg','jpeg','gif','webp'}
for d in (MEDIA_FOLDER, PROFILE_FOLDER, MEETING_FOLDER):
    os.makedirs(d, exist_ok=True)

VENUES = [
    "5th Floor Lecture Hall",
    "4th Floor Lecture Hall A",
    "4th Floor Lecture Hall B",
    "3rd Floor Lecture Hall A",
    "3rd Floor Lecture Hall B",
    "2nd Floor Computer Lab",
    "2nd Floor Bio-Lab",
    "Basement Research Room"
]
DAYS = ['MONDAY','TUESDAY','WEDNESDAY','THURSDAY','FRIDAY','SATURDAY','SUNDAY']

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80)); ip = s.getsockname()[0]; s.close(); return ip
    except: return '127.0.0.1'

# ── Generic JSON helpers ─────────────────────────────────────
def _load(path, default):
    if not os.path.exists(path):
        with open(path,'w') as f: json.dump(default, f, indent=2)
        return default
    with open(path) as f: return json.load(f)

def _save(path, obj):
    with open(path,'w') as f: json.dump(obj, f, indent=2)

# ── Timetable data ───────────────────────────────────────────
def empty_week():
    return {day: {v: {"sessions": []} for v in VENUES} for day in DAYS}

def load_data():
    if not os.path.exists(DATA_FILE):
        d = {"active_day": "MONDAY", "week": empty_week(), "media": [],
             "settings": {"slide_interval": 8, "timetable_duration": 30}}
        save_data(d); return d
    return _load(DATA_FILE, {})

def save_data(d): _save(DATA_FILE, d)

# ── Presets (kept for backward-compat autocomplete) ──────────
def load_presets():
    return _load(PRESETS_FILE, {
        "programs":  ["DCS 01","DCS 02","BCS 01","BCS 02","DHTM 01","DHTM 02","DBMS 01","DBMS 02","PhD Sessions"],
        "subjects":  ["Introduction to Accounting","Information Technology","Data Structures",
                      "Object Oriented Programming","Database Management","Web Development"],
        "lecturers": ["Ms. Janani Perera","Mr. Imal Perera","Mr. Uditha Lashan"]
    })
def save_presets(p): _save(PRESETS_FILE, p)

# ── Catalog: programs -> modules -> lessons ──────────────────
def load_catalog():
    return _load(CATALOG_FILE, {"programs": []})
    # structure:
    # { "programs": [
    #     { "name":"DCS 01", "modules":[
    #         { "name":"Object Oriented Programming", "code":"CCS20704",
    #           "lessons":["Introduction to OOP","Classes & Objects","Inheritance"] }
    #     ]}
    # ]}
def save_catalog(c): _save(CATALOG_FILE, c)

# ── Lecturers with profile images ───────────────────────────
def load_lecturers():
    return _load(LECTURERS_FILE, {"lecturers": []})
    # { "lecturers":[ {"id":"..","name":"Mr. Uditha Lashan","image":"abc.jpg"} ] }
def save_lecturers(l): _save(LECTURERS_FILE, l)

# ── Templates ────────────────────────────────────────────────
def load_templates(): return _load(TEMPLATES_FILE, {})
def save_templates(t): _save(TEMPLATES_FILE, t)

# ── Special meeting ──────────────────────────────────────────
def load_meeting():
    return _load(MEETING_FILE, {"active": False, "title": "", "text": "", "media": []})
def save_meeting(m): _save(MEETING_FILE, m)

# ════════════════════════════════════════════════════════════
#  API
# ════════════════════════════════════════════════════════════
@app.route('/api/data')
def get_data():
    d = load_data()
    d['meeting'] = load_meeting()
    return jsonify(d)

@app.route('/api/network-info')
def network_info():
    ip = get_local_ip()
    return jsonify({'ip': ip, 'admin': f'http://{ip}:5000', 'display': f'http://{ip}:5000/display'})

@app.route('/api/firebase-config')
def firebase_config():
    cfg = {
        'apiKey':            os.environ.get('FIREBASE_API_KEY',''),
        'authDomain':        os.environ.get('FIREBASE_AUTH_DOMAIN',''),
        'projectId':         os.environ.get('FIREBASE_PROJECT_ID',''),
        'storageBucket':     os.environ.get('FIREBASE_STORAGE_BUCKET',''),
        'messagingSenderId': os.environ.get('FIREBASE_MESSAGING_SENDER_ID',''),
        'appId':             os.environ.get('FIREBASE_APP_ID',''),
    }
    configured = bool(cfg['apiKey']) and cfg['apiKey'] != 'PASTE_YOUR_API_KEY'
    return jsonify({'configured': configured, 'config': cfg if configured else None})

# ── Presets ──────────────────────────────────────────────────
@app.route('/api/presets')
def get_presets(): return jsonify(load_presets())

@app.route('/api/presets', methods=['POST'])
def update_presets():
    p = load_presets(); body = request.json
    if '_replace' in body:
        for k,v in body['_replace'].items(): p[k] = v
    else:
        for key in ('programs','subjects','lecturers'):
            if key in body:
                val = body[key].strip()
                if val and val not in p.get(key,[]): p.setdefault(key,[]).append(val)
    save_presets(p); return jsonify({'ok':True})

# ── Catalog (programs/modules/lessons) ──────────────────────
@app.route('/api/catalog')
def get_catalog(): return jsonify(load_catalog())

@app.route('/api/catalog', methods=['POST'])
def save_catalog_route():
    body = request.json
    if 'programs' in body:
        save_catalog({'programs': body['programs']})
        return jsonify({'ok':True})
    return jsonify({'error':'Invalid'}),400

# ── Lecturers ────────────────────────────────────────────────
@app.route('/api/lecturers')
def get_lecturers(): return jsonify(load_lecturers())

@app.route('/api/lecturers', methods=['POST'])
def add_lecturer():
    data = load_lecturers()
    body = request.json
    name = (body.get('name') or '').strip()
    if not name: return jsonify({'error':'Name required'}),400
    # update existing or add new
    existing = next((l for l in data['lecturers'] if l['name']==name), None)
    if existing:
        if 'image' in body: existing['image'] = body['image']
    else:
        data['lecturers'].append({'id':uuid.uuid4().hex,'name':name,'image':body.get('image','')})
    save_lecturers(data)
    return jsonify({'ok':True})

@app.route('/api/lecturers/delete/<lid>', methods=['DELETE'])
def delete_lecturer(lid):
    data = load_lecturers()
    item = next((l for l in data['lecturers'] if l['id']==lid),None)
    if item and item.get('image'):
        fp = os.path.join(PROFILE_FOLDER, item['image'])
        if os.path.exists(fp): os.remove(fp)
    data['lecturers'] = [l for l in data['lecturers'] if l['id']!=lid]
    save_lecturers(data)
    return jsonify({'ok':True})

@app.route('/api/lecturers/upload-image', methods=['POST'])
def upload_profile():
    if 'file' not in request.files: return jsonify({'error':'No file'}),400
    file = request.files['file']
    if file and '.' in file.filename and file.filename.rsplit('.',1)[1].lower() in IMG_EXT:
        ext = file.filename.rsplit('.',1)[1].lower()
        fname = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(PROFILE_FOLDER, fname))
        return jsonify({'ok':True,'filename':fname})
    return jsonify({'error':'Image files only'}),400

# ── Excel import ─────────────────────────────────────────────
@app.route('/api/import-excel', methods=['POST'])
def import_excel():
    """
    Expects an .xlsx with sheets (any subset):
      - 'Programs'  : column A = program name
      - 'Modules'   : A=program, B=module name, C=module code(optional)
      - 'Lessons'   : A=program, B=module, C=lesson
      - 'Lecturers' : A=lecturer name
    Falls back to a single sheet with columns: Program | Module | Code | Lesson | Lecturer
    """
    if 'file' not in request.files: return jsonify({'error':'No file'}),400
    file = request.files['file']
    if not (file and file.filename.lower().endswith(('.xlsx','.xls'))):
        return jsonify({'error':'Excel files only (.xlsx)'}),400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'error':f'Could not read Excel: {e}'}),400

    catalog = load_catalog()
    presets = load_presets()
    lecturers = load_lecturers()

    def get_program(name):
        name = str(name).strip()
        prog = next((p for p in catalog['programs'] if p['name']==name), None)
        if not prog:
            prog = {'name':name,'modules':[]}
            catalog['programs'].append(prog)
            if name not in presets.setdefault('programs',[]): presets['programs'].append(name)
        return prog

    def get_module(prog, mname, mcode=''):
        mname = str(mname).strip()
        mod = next((m for m in prog['modules'] if m['name']==mname), None)
        if not mod:
            mod = {'name':mname,'code':str(mcode or '').strip(),'lessons':[]}
            prog['modules'].append(mod)
            if mname not in presets.setdefault('subjects',[]): presets['subjects'].append(mname)
        elif mcode and not mod.get('code'):
            mod['code'] = str(mcode).strip()
        return mod

    sheets = {s.lower(): s for s in wb.sheetnames}
    counts = {'programs':0,'modules':0,'lessons':0,'lecturers':0}

    def rows_of(sheet_name):
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(c is not None and str(c).strip() for c in row):
                yield row

    # Dedicated sheets
    if 'programs' in sheets:
        for row in rows_of(sheets['programs']):
            if row[0]: get_program(row[0]); counts['programs']+=1
    if 'modules' in sheets:
        for row in rows_of(sheets['modules']):
            if row[0] and len(row)>1 and row[1]:
                prog = get_program(row[0])
                get_module(prog, row[1], row[2] if len(row)>2 else '')
                counts['modules']+=1
    if 'lessons' in sheets:
        for row in rows_of(sheets['lessons']):
            if row[0] and len(row)>2 and row[1] and row[2]:
                prog = get_program(row[0]); mod = get_module(prog, row[1])
                lesson = str(row[2]).strip()
                if lesson not in mod['lessons']: mod['lessons'].append(lesson); counts['lessons']+=1
    if 'lecturers' in sheets:
        for row in rows_of(sheets['lecturers']):
            if row[0]:
                lname = str(row[0]).strip()
                if not any(l['name']==lname for l in lecturers['lecturers']):
                    lecturers['lecturers'].append({'id':uuid.uuid4().hex,'name':lname,'image':''})
                if lname not in presets.setdefault('lecturers',[]): presets['lecturers'].append(lname)
                counts['lecturers']+=1

    # Single combined sheet fallback (if no dedicated sheets matched)
    if sum(counts.values())==0:
        ws = wb[wb.sheetnames[0]]
        header = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
        def col(name):
            for i,h in enumerate(header):
                if name in h: return i
            return -1
        ci = {k:col(k) for k in ['program','module','code','lesson','lecturer']}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            pv = row[ci['program']] if ci['program']>=0 and ci['program']<len(row) else None
            if pv:
                prog = get_program(pv); counts['programs']+=1
                mv = row[ci['module']] if ci['module']>=0 and ci['module']<len(row) else None
                if mv:
                    cv = row[ci['code']] if ci['code']>=0 and ci['code']<len(row) else ''
                    mod = get_module(prog, mv, cv); counts['modules']+=1
                    lv = row[ci['lesson']] if ci['lesson']>=0 and ci['lesson']<len(row) else None
                    if lv:
                        lv=str(lv).strip()
                        if lv not in mod['lessons']: mod['lessons'].append(lv); counts['lessons']+=1
            lecv = row[ci['lecturer']] if ci['lecturer']>=0 and ci['lecturer']<len(row) else None
            if lecv:
                lname=str(lecv).strip()
                if not any(l['name']==lname for l in lecturers['lecturers']):
                    lecturers['lecturers'].append({'id':uuid.uuid4().hex,'name':lname,'image':''})
                if lname not in presets.setdefault('lecturers',[]): presets['lecturers'].append(lname)
                counts['lecturers']+=1

    save_catalog(catalog); save_presets(presets); save_lecturers(lecturers)
    return jsonify({'ok':True,'imported':counts})

# ── Timetable ────────────────────────────────────────────────
@app.route('/api/timetable', methods=['POST'])
def update_timetable():
    data = load_data(); body = request.json
    day  = body.get('day', data.get('active_day','MONDAY'))
    rows = body.get('rows', {})
    if 'week' not in data: data['week'] = empty_week()
    data['week'][day] = rows
    data['active_day'] = day
    p = load_presets(); changed = False
    for venue, vdata in rows.items():
        for sess in vdata.get('sessions', []):
            for field, pkey in [('program','programs'),('subject','subjects'),('lecturer','lecturers')]:
                val = (sess.get(field) or '').strip()
                if val and val not in p.get(pkey,[]): p.setdefault(pkey,[]).append(val); changed = True
    if changed: save_presets(p)
    save_data(data); return jsonify({'ok':True})

@app.route('/api/active-day', methods=['POST'])
def set_active_day():
    data = load_data()
    day = request.json.get('day','MONDAY')
    if day in DAYS: data['active_day'] = day
    save_data(data); return jsonify({'ok':True})

# ── Templates ────────────────────────────────────────────────
@app.route('/api/templates')
def get_templates(): return jsonify(load_templates())

@app.route('/api/templates/save-week', methods=['POST'])
def save_week_template():
    body = request.json; name = (body.get('name') or '').strip()
    if not name: return jsonify({'error':'Name required'}),400
    data = load_data(); templates = load_templates()
    templates[name] = {'week': data.get('week', empty_week()),
                       'saved': datetime.now().strftime('%Y-%m-%d %H:%M')}
    save_templates(templates); return jsonify({'ok':True, 'name':name})

@app.route('/api/templates/load/<name>')
def load_template(name):
    templates = load_templates()
    if name not in templates: return jsonify({'error':'Not found'}),404
    return jsonify(templates[name])

@app.route('/api/templates/delete/<name>', methods=['DELETE'])
def delete_template(name):
    templates = load_templates()
    if name in templates: del templates[name]; save_templates(templates)
    return jsonify({'ok':True})

# ── Media (flyers/slideshow) ────────────────────────────────
@app.route('/api/media/upload', methods=['POST'])
def upload_media():
    if 'file' not in request.files: return jsonify({'error':'No file'}),400
    file = request.files['file']
    if file and '.' in file.filename and file.filename.rsplit('.',1)[1].lower() in ALLOWED_EXT:
        ext  = file.filename.rsplit('.',1)[1].lower()
        fname= f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(MEDIA_FOLDER, fname))
        data = load_data()
        data['media'].append({'id':uuid.uuid4().hex,'filename':fname,
            'original':secure_filename(file.filename),
            'type':'video' if ext in {'mp4','webm','mov'} else 'image',
            'added':datetime.now().isoformat()})
        save_data(data); return jsonify({'ok':True,'filename':fname})
    return jsonify({'error':'Not allowed'}),400

@app.route('/api/media/delete/<mid>', methods=['DELETE'])
def delete_media(mid):
    data = load_data()
    item = next((m for m in data['media'] if m['id']==mid),None)
    if item:
        fp = os.path.join(MEDIA_FOLDER, item['filename'])
        if os.path.exists(fp): os.remove(fp)
        data['media'] = [m for m in data['media'] if m['id']!=mid]
        save_data(data)
    return jsonify({'ok':True})

# ── Special Meeting ──────────────────────────────────────────
@app.route('/api/meeting')
def get_meeting(): return jsonify(load_meeting())

@app.route('/api/meeting', methods=['POST'])
def set_meeting():
    body = request.json
    m = load_meeting()
    m['active'] = bool(body.get('active', m.get('active', False)))
    if 'title' in body: m['title'] = body['title']
    if 'text'  in body: m['text']  = body['text']
    if 'media' in body: m['media'] = body['media']
    save_meeting(m); return jsonify({'ok':True})

@app.route('/api/meeting/upload', methods=['POST'])
def meeting_upload():
    if 'file' not in request.files: return jsonify({'error':'No file'}),400
    file = request.files['file']
    if file and '.' in file.filename and file.filename.rsplit('.',1)[1].lower() in ALLOWED_EXT:
        ext = file.filename.rsplit('.',1)[1].lower()
        fname = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(MEETING_FOLDER, fname))
        return jsonify({'ok':True,'filename':fname,
                        'type':'video' if ext in {'mp4','webm','mov'} else 'image'})
    return jsonify({'error':'Not allowed'}),400

@app.route('/api/meeting/clear', methods=['POST'])
def meeting_clear():
    m = load_meeting()
    for item in m.get('media',[]):
        fp = os.path.join(MEETING_FOLDER, item.get('filename',''))
        if os.path.exists(fp):
            try: os.remove(fp)
            except: pass
    save_meeting({"active": False, "title": "", "text": "", "media": []})
    return jsonify({'ok':True})

@app.route('/api/settings', methods=['POST'])
def update_settings():
    data = load_data(); data['settings'].update(request.json); save_data(data); return jsonify({'ok':True})

# ── Static ───────────────────────────────────────────────────
@app.route('/media/<path:filename>')
def serve_media(filename): return send_from_directory(MEDIA_FOLDER, filename)

@app.route('/profiles/<path:filename>')
def serve_profile(filename): return send_from_directory(PROFILE_FOLDER, filename)

@app.route('/meeting-media/<path:filename>')
def serve_meeting_media(filename): return send_from_directory(MEETING_FOLDER, filename)

@app.route('/logos/<path:filename>')
def serve_logos(filename): return send_from_directory('public', filename)

@app.route('/')
def admin(): return send_from_directory('.', 'admin.html')

@app.route('/login')
def login(): return send_from_directory('.', 'login.html')

@app.route('/display')
def display(): return send_from_directory('.', 'display.html')

if __name__ == '__main__':
    ip = get_local_ip()
    print("\n" + "="*56)
    print("  MSI Campus Digital Signage")
    print("="*56)
    print(f"  Login / Admin : http://localhost:5000")
    print(f"  TV Display    : http://{ip}:5000/display")
    print("="*56 + "\n")
    app.run(host='0.0.0.0', port=5000, debug=False)
